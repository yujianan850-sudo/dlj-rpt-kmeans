#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Incremental KMeans retrain from device_model to device_model_new.

Rules:
  - Iterate IMEIs from device_model and device_model_new (union)
  - Missing in device_model_new: first train with latest WINDOW_DAYS data
  - Existing in device_model_new: keep saved pen, append after state_doc.end
  - Train success -> update device_model_new
  - Train fail -> keep device_model_new unchanged
  - Export excel summary with retrain date suffix
"""

import datetime
import json
import os
import pickle
import sys
import threading
import time
import warnings
from concurrent.futures import ThreadPoolExecutor

import numpy as np
import pandas as pd
import pymongo
import ruptures as rpt_lib
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy.stats import kurtosis, skew
from sklearn.cluster import KMeans
from sklearn.metrics import silhouette_score

warnings.filterwarnings("ignore")

# ============================================================
#  Config
# ============================================================

# First-train window. Incremental runs use state_doc.end -> run_end instead.
WINDOW_DAYS = int(os.environ.get("WINDOW_DAYS", "15"))

OUTPUT_DIR = os.environ.get("OUTPUT_DIR", ".")
LOG_DIR = os.environ.get("LOG_DIR", "logs")
RETRAIN_INTERVAL_DAYS = int(os.environ.get("RETRAIN_INTERVAL_DAYS", "7"))
MAX_TRAIN_COUNT = int(os.environ.get("MAX_TRAIN_COUNT", "-1"))  # -1 means train all
WORKER_THREADS = int(os.environ.get("WORKER_THREADS", "5"))
# legacy cap constant kept for compatibility with helper functions
FEATURE_MATRIX_MAX_ROWS = int(os.environ.get("FEATURE_MATRIX_MAX_ROWS", "120000"))
_u_wave_env = os.environ.get("TRAIN_U_WAVE_OVERRIDE", "0.05").strip()
TRAIN_U_WAVE_OVERRIDE = None if _u_wave_env.lower() in ("", "none", "null") else float(_u_wave_env)
_run_end_env = os.environ.get("RUN_END_DATE_OVERRIDE", "").strip()
RUN_END_DATE_OVERRIDE = _run_end_env or None

# MongoDB
MONGO_HOST = os.environ.get("MONGO_HOST", "172.22.22.120")
MONGO_PORT = int(os.environ.get("MONGO_PORT", "27017"))
MONGO_USER = os.environ.get("MONGO_USER", "guiwu")
MONGO_PWD = os.environ.get("MONGO_PWD", "106ling106")
MONGO_AUTH_DB = os.environ.get("MONGO_AUTH_DB", "admin")
MONGO_TARGET_DB = os.environ.get("MONGO_TARGET_DB", "galv-center")

# device_model: bootstrap device list + first-train source (read-only, not updated by this script)
BASE_COLLECTION  = "device_model"
# Retrained models destination
STATE_COLLECTION = "device_model_new"

# Training params (same as production)
TRAIN_LENGTH      = 300
TRAIN_PEN_LIST    = [20, 10, 5, 1, 0.1, 0.01, 0.001]
TRAIN_FEATURE_LIST = ["mean", "max_value", "min_value"]

# ============================================================
#  DB
# ============================================================

def _get_mongo_db():
    uri = "mongodb://{}:{}@{}:{}/{}".format(
        MONGO_USER, MONGO_PWD, MONGO_HOST, MONGO_PORT, MONGO_AUTH_DB
    )
    client = pymongo.MongoClient(uri, maxPoolSize=50, serverSelectionTimeoutMS=10000)
    return client, client[MONGO_TARGET_DB]


def fetch_doc_by_imei(db, table_name, imei):
    candidates = [imei]
    if isinstance(imei, str) and imei.isdigit():
        candidates.append(int(imei))
    for c in candidates:
        doc = db[table_name].find_one({"imei": c})
        if doc:
            return doc
    return None


def _fetch_current_entries(db, collection_name, imei, start_ms, end_ms):
    pipeline = [
        {"$match": {"imei": imei}},
        {"$unwind": "$dataList"},
        {"$match": {"dataList.time": {"$gte": start_ms, "$lt": end_ms}}},
        {"$project": {"_id": 0, "dataList": 1}},
    ]
    try:
        return list(db[collection_name].aggregate(pipeline))
    except Exception as e:
        print("      Mongo error on {}: {}".format(collection_name, e))
        return False


def fetch_train_data(db, imei, start_ms, end_ms):
    all_results = []
    for collection_name in ("device_current", "to_archived_current"):
        results = _fetch_current_entries(db, collection_name, imei, start_ms, end_ms)
        if results is False:
            return False
        all_results.extend(results)

    all_results.sort(key=lambda item: item.get("dataList", {}).get("time", 0))

    processed = []
    seen_times = set()
    for entry in all_results:
        dl = entry.get("dataList", {})
        try:
            value   = float(dl["value"])
            minutes = int(dl["minutes"])
            data_time = int(dl["time"])
        except (KeyError, ValueError, TypeError):
            continue
        if data_time in seen_times:
            continue
        seen_times.add(data_time)
        if minutes < 0:
            continue
        processed.extend([value] * minutes)

    return processed if processed else False


def save_to_state(db, imei, doc, collection_name=STATE_COLLECTION):
    doc = {k: v for k, v in doc.items() if k != "_id"}
    imei_str = str(imei)
    doc["imei"] = imei_str
    if imei_str.isdigit():
        db[collection_name].delete_many({"imei": int(imei_str)})
    db[collection_name].update_one(
        {"imei": imei_str},
        {"$set": doc},
        upsert=True,
    )


# ============================================================
#  Time utils
# ============================================================

def make_time_range_ms(start_days_back, end_days_back):
    today    = datetime.datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    start_ms = int((today - datetime.timedelta(days=start_days_back)).timestamp() * 1000)
    end_ms   = int((today - datetime.timedelta(days=end_days_back)).timestamp() * 1000)
    return start_ms, end_ms


def get_run_end_ms():
    if RUN_END_DATE_OVERRIDE is None:
        return make_time_range_ms(0, 0)[1]
    dt = datetime.datetime.strptime(RUN_END_DATE_OVERRIDE, "%Y-%m-%d")
    dt = dt.replace(hour=0, minute=0, second=0, microsecond=0)
    return int(dt.timestamp() * 1000)


def make_first_train_range_ms(run_end_ms=None):
    if run_end_ms is None:
        return make_time_range_ms(WINDOW_DAYS, 0)
    start_ms = int(run_end_ms - WINDOW_DAYS * 24 * 60 * 60 * 1000)
    return start_ms, int(run_end_ms)


def make_incremental_range_ms(last_end_ms, run_end_ms=None):
    today_end_ms = make_time_range_ms(0, 0)[1] if run_end_ms is None else int(run_end_ms)
    if last_end_ms is None:
        return make_first_train_range_ms(today_end_ms)
    return int(last_end_ms), today_end_ms


def ts_to_str(ts_ms):
    if ts_ms is None:
        return ""
    try:
        return datetime.datetime.fromtimestamp(ts_ms / 1000).strftime("%Y-%m-%d")
    except Exception:
        return str(ts_ms)


def resolve_output_filepath(run_end_ms):
    """Build excel path with retrain date suffix under OUTPUT_DIR."""
    date_suffix = datetime.datetime.fromtimestamp(run_end_ms / 1000).strftime("%Y%m%d")
    return os.path.join(OUTPUT_DIR, "incremental_train_{}.xlsx".format(date_suffix))


def load_primary_device_list(db):
    """
    Load IMEIs from device_model and device_model_new.
    device_model is the bootstrap source; device_model_new-only IMEIs are still retrained.
    """
    seen = {}
    for doc in db[BASE_COLLECTION].find({}, {"_id": 0, "imei": 1}):
        imei = str(doc.get("imei", "")).strip()
        if imei:
            seen[imei] = {"imei": imei}
    for doc in db[STATE_COLLECTION].find({}, {"_id": 0, "imei": 1}):
        imei = str(doc.get("imei", "")).strip()
        if imei and imei not in seen:
            seen[imei] = {"imei": imei}
    return [seen[k] for k in sorted(seen.keys())]


# ============================================================
#  Feature extraction
# ============================================================

def _rms(arr):
    return np.sqrt(np.mean(np.square(arr)))


def _get_features(arr, feature_list):
    feature_fn = {
        "mean":         np.mean,
        "var":          np.var,
        "skew":         skew,
        "kurtosis":     kurtosis,
        "max_value":    np.max,
        "min_value":    np.min,
        "peak_to_peak": np.ptp,
        "median_value": np.median,
        "rms":          _rms,
    }
    features = []
    for f in feature_list:
        result = feature_fn[f](arr)
        if isinstance(result, np.ndarray):
            features.extend(result)
        else:
            features.append(result)
    return np.array(features)


def _sequence_processing(seg_data, feature_list, U_close):
    padded_seqs, seq_pos = [], []
    for i, seq in enumerate(seg_data):
        if len(seq) == 0:
            continue
        mean_val = np.mean(seq)
        if mean_val < 0 or mean_val < U_close:
            continue
        padded_seqs.append(_get_features(seq, feature_list))
        seq_pos.append(i)
    return np.array(padded_seqs), seq_pos


# ============================================================
#  Segmentation
# ============================================================

def _divide_by_negative(data, base_idx):
    result, child_arr, negative_arr = [], [], []
    i = 0
    for i, e in enumerate(data):
        if e != -1:
            child_arr.append(e)
            if negative_arr:
                result.append((negative_arr, base_idx + i))
                negative_arr = []
        else:
            negative_arr.append(e)
            if child_arr:
                result.append((child_arr, base_idx + i))
                child_arr = []
    if child_arr:
        result.append((child_arr, base_idx + len(child_arr) + i))
    if negative_arr:
        result.append((negative_arr, base_idx + len(negative_arr) + i))
    return result


def _split_sequence(seq, length):
    return [(seq[i: i + length], i // length) for i in range(0, len(seq), length)]


def _generate_subsequences(arr, point_list):
    seg_arr, start = [], 0
    for point in point_list:
        if start == point:
            continue
        seg_arr.append(arr[start:point])
        start = point
    return seg_arr


def _generate_seglist(data, length, pen):
    sequence_list = _split_sequence(data, length)
    point_list = []
    for seg in sequence_list:
        first_seq_data = _divide_by_negative(seg[0], seg[1] * length)
        for i, e in enumerate(first_seq_data):
            if len(e[0]) == 1:
                point_list.append(e[1])
                continue
            algo = rpt_lib.Binseg(model="rbf").fit(np.array(e[0]))
            change_point = algo.predict(pen=pen)
            base_point = seg[1] * length if i == 0 else first_seq_data[i - 1][1]
            for cp in change_point:
                point_list.append(cp + base_point)
    return _generate_subsequences(data, point_list)


# ============================================================
#  Extract feature matrix from raw data for one pen
# ============================================================

def extract_features(raw_data, pen, U_close):
    """
    Returns np.ndarray (N x n_features) or None if insufficient data.
    """
    try:
        seg_data = _generate_seglist(raw_data, TRAIN_LENGTH, pen)
    except Exception:
        return None
    feat, _ = _sequence_processing(seg_data, TRAIN_FEATURE_LIST, U_close)
    return feat if len(feat) > 0 else None


# ============================================================
#  KMeans on combined feature matrix
# ============================================================

def _judge_wave(feat, U_wave):
    diff = round(float(np.max(feat[:, 0]) - np.min(feat[:, 0])), 2)
    return diff >= U_wave


def train_on_features(feat_combined, U_wave):
    """
    Run KMeans(k=2) on feat_combined, select best pen by ss_score.
    Returns (best_result_dict, km_object) or (None, None).

    best_result_dict keys: pen, ss_score, mean_seg
    Note: pen here refers to which pen produced feat_combined;
    since combined matrix already exists, we just cluster it once.
    """
    if len(feat_combined) <= 3:
        return None, None
    if not _judge_wave(feat_combined, U_wave):
        return None, None

    km     = KMeans(n_clusters=2, n_init=10, random_state=42)
    labels = km.fit_predict(feat_combined)
    ss     = float(silhouette_score(feat_combined, labels))

    mean_seg = [round(float(center[0]), 4) for center in km.cluster_centers_]
    result = {"ss_score": round(ss, 4), "mean_seg": mean_seg}
    return result, km


def incremental_retrain(old_feat, raw_new, pen, U_close, U_wave):
    """
    For a single pen:
      1. Extract features from raw_new (may be None if no new data)
      2. Stack with old_feat (may be None if first run)
      3. Train KMeans on combined matrix
    Returns (result_dict, km_object, combined_feat) or (None, None, None)
    """
    new_feat = None
    if raw_new is not False and len(raw_new) > 0:
        new_feat = extract_features(raw_new, pen, U_close)

    if old_feat is None and new_feat is None:
        return None, None, None
    elif old_feat is None:
        combined = new_feat
    elif new_feat is None:
        combined = old_feat
    else:
        combined = np.vstack([old_feat, new_feat])

    result, km = train_on_features(combined, U_wave)
    return result, km, combined


# ============================================================
#  Best pen selection across all pens
# ============================================================

def best_incremental(old_feat, raw_new, U_close, U_wave, new_start_ms, new_end_ms):
    """
    Try all pens, pick the one with highest ss_score.
    Returns (best_result, best_km, best_combined_feat) or (None, None, None).
    best_result includes 'pen' field.
    """
    best_score   = -np.inf
    best_result  = None
    best_km      = None
    best_combined = None

    for pen in TRAIN_PEN_LIST:
        result, km, combined = incremental_retrain(old_feat, raw_new, pen, U_close, U_wave)
        if result is None:
            continue
        if result["ss_score"] > best_score:
            best_score    = result["ss_score"]
            best_result   = dict(result)
            best_result["pen"] = pen
            best_km       = km
            best_combined = combined

    return best_result, best_km, best_combined


# ============================================================
#  feature_matrix size guard (MongoDB 16MB doc limit)
# ============================================================

def _estimate_matrix_json_bytes(matrix_list):
    try:
        return len(json.dumps(matrix_list, separators=(",", ":")).encode("utf-8"))
    except Exception:
        return sys.getsizeof(matrix_list)


def trim_feature_rows(feat, max_rows):
    """Keep the most recent max_rows rows. Returns (trimmed, n_removed)."""
    if feat is None or len(feat) <= max_rows:
        return feat, 0
    removed = len(feat) - max_rows
    return feat[-max_rows:].copy(), removed


def apply_feature_cap(feat, best_result, best_km, U_wave, log_print=print):
    """
    If feat exceeds FEATURE_MATRIX_MAX_ROWS, trim to tail and re-cluster
    so mean_seg/ss_score match the stored matrix. pen is unchanged.
    Returns (feat_out, result_out, km_out).
    """
    feat2, removed = trim_feature_rows(feat, FEATURE_MATRIX_MAX_ROWS)
    if removed == 0:
        return feat2, best_result, best_km
    log_print("    [CAP] feature_matrix trimmed by {} rows -> {}".format(removed, len(feat2)))
    r2, km2 = train_on_features(feat2, U_wave)
    if r2 is None:
        return feat2, best_result, best_km
    new_res = {
        "pen":      best_result["pen"],
        "ss_score": r2["ss_score"],
        "mean_seg": r2["mean_seg"],
    }
    log_print("    [CAP] re-cluster on trimmed: ss={} mean_seg={}".format(
        new_res["ss_score"], new_res["mean_seg"]))
    return feat2, new_res, km2


# ============================================================
#  State merge + per-device worker
# ============================================================

def _build_excel_row(imei, mode, status, old_mean_seg, best_result, old_range="", new_range=""):
    old_mean_seg_str = ""
    if old_mean_seg:
        old_mean_seg_str = str(old_mean_seg)
    new_mean_seg_str = ""
    if best_result is not None:
        ms   = best_result.get("mean_seg", [])
        new_mean_seg_str = str(ms)
    return {
        "IMEI": imei,
        "mode": mode,
        "status": status,
        "old_train_range": old_range,
        "old_mean_seg": old_mean_seg_str,
        "new_train_range": new_range,
        "new_mean_seg": new_mean_seg_str,
    }


def process_device(work):
    """
    work = (idx, total, base_doc, mongo_db, tprint, run_end_ms)
    First run trains on latest WINDOW_DAYS data.
    Later runs append only unseen data with the saved pen.
    """
    idx, total, base_item, mongo_db, tprint, run_end_ms = work
    imei = str(base_item.get("imei", ""))
    base_doc = None
    state_doc = fetch_doc_by_imei(mongo_db, STATE_COLLECTION, imei)
    source_doc = dict(state_doc) if state_doc is not None else {}
    if state_doc is None:
        base_doc = fetch_doc_by_imei(mongo_db, BASE_COLLECTION, imei)
        if base_doc is None:
            tprint("    missing base doc -> skip")
            return {
                "idx": idx, "outcome": "keep_state",
                "row": _build_excel_row(imei, "missing_base_doc", "keep_old(no_base_doc)", [], None, "", "keep_old(no_base_doc)"),
                "counters": {"keep_state": 1},
            }
        source_doc = dict(base_doc)
    has_new_model = state_doc is not None and state_doc.get("feature_matrix") is not None

    U_close = float(source_doc.get("U_close", 0.05))
    if TRAIN_U_WAVE_OVERRIDE is None:
        U_wave = float(source_doc.get("U_wave", 0.10))
    else:
        U_wave = float(TRAIN_U_WAVE_OVERRIDE)

    try:
        old_mean_seg = sorted([round(float(x), 4) for x in source_doc.get("mean_seg", [])])
    except Exception:
        old_mean_seg = []

    if has_new_model:
        start_ms, end_ms = make_incremental_range_ms(state_doc.get("end"), run_end_ms)
        mode = "incremental_fixed_pen"
    else:
        start_ms, end_ms = make_first_train_range_ms(run_end_ms)
        mode = "first_train_from_state" if state_doc is not None else "first_train_from_base"
    tprint("  [{:04d}/{}] IMEI: {}  [{}]".format(idx, total, imei, mode))
    tprint("    train_range={} ~ {}".format(ts_to_str(start_ms), ts_to_str(end_ms)))

    raw_data = fetch_train_data(mongo_db, imei, start_ms, end_ms)
    old_range = "{} ~ {}".format(ts_to_str(source_doc.get("start")), ts_to_str(source_doc.get("end")))
    model_start_ms = source_doc.get("start") if has_new_model else start_ms
    current_range = "{} ~ {}".format(ts_to_str(model_start_ms), ts_to_str(end_ms))

    if raw_data is False:
        if state_doc is not None:
            tprint("    no data -> keep existing state (no overwrite)")
            return {
                "idx": idx, "outcome": "keep_state",
                "row": _build_excel_row(imei, mode, "keep_old(no_data)", old_mean_seg, None, old_range, "keep_old(no_data)"),
                "counters": {"keep_state": 1},
            }
        tprint("    no data -> seed from old model once")
        out_doc = dict(base_doc)
        out_doc["U_wave"] = U_wave
        save_to_state(mongo_db, imei, out_doc, collection_name=STATE_COLLECTION)
        return {
            "idx": idx, "outcome": "seed_old_once",
            "row": _build_excel_row(imei, mode, "seed_old(no_data)", old_mean_seg, None, old_range, "seed_old(no_data)"),
            "counters": {"seed_old_once": 1},
        }

    tprint("    data={}min".format(len(raw_data)))

    old_feat = None
    if has_new_model:
        try:
            old_feat = np.array(state_doc.get("feature_matrix"), dtype=float)
            if len(old_feat) == 0:
                old_feat = None
        except Exception:
            old_feat = None
            has_new_model = False

    best_score = -np.inf
    best_result = None
    best_km = None
    best_pen = None
    best_combined = None
    if has_new_model:
        fixed_pen = state_doc.get("pen")
        if fixed_pen is None:
            tprint("    missing saved pen -> keep existing state (no overwrite)")
            return {
                "idx": idx, "outcome": "keep_state",
                "row": _build_excel_row(imei, mode, "keep_old(missing_pen)", old_mean_seg, None, old_range, "keep_old(missing_pen)"),
                "counters": {"keep_state": 1},
            }
        new_feat = extract_features(raw_data, fixed_pen, U_close)
        if new_feat is None or len(new_feat) == 0:
            tprint("    no usable new features under fixed pen -> keep existing state")
            return {
                "idx": idx, "outcome": "keep_state",
                "row": _build_excel_row(imei, mode, "keep_old(no_new_feat)", old_mean_seg, None, old_range, "keep_old(no_new_feat)"),
                "counters": {"keep_state": 1},
            }
        combined = np.vstack([old_feat, new_feat]) if old_feat is not None else new_feat
        result, km = train_on_features(combined, U_wave)
        if result is not None:
            best_score = result["ss_score"]
            best_result = result
            best_km = km
            best_pen = fixed_pen
            best_combined = combined
    else:
        for pen in TRAIN_PEN_LIST:
            result, km, combined = incremental_retrain(old_feat, raw_data, pen, U_close, U_wave)
            if result is None:
                continue
            if result["ss_score"] > best_score:
                best_score = result["ss_score"]
                best_result = result
                best_km = km
                best_pen = pen
                best_combined = combined

    if best_result is None:
        if state_doc is not None:
            tprint("    train failed -> keep existing state (no overwrite)")
            return {
                "idx": idx, "outcome": "keep_state",
                "row": _build_excel_row(imei, mode, "keep_old(train_fail)", old_mean_seg, None, old_range, "keep_old(train_fail)"),
                "counters": {"keep_state": 1},
            }
        tprint("    train failed -> seed from old model once")
        out_doc = dict(base_doc)
        out_doc["U_wave"] = U_wave
        save_to_state(mongo_db, imei, out_doc, collection_name=STATE_COLLECTION)
        return {
            "idx": idx, "outcome": "seed_old_once",
            "row": _build_excel_row(imei, mode, "seed_old(train_fail)", old_mean_seg, None, old_range, "seed_old(train_fail)"),
            "counters": {"seed_old_once": 1},
        }

    tprint("    -> OK pen={} ss={} mean_seg={}".format(
        best_pen, best_result["ss_score"], best_result["mean_seg"]))

    best_combined, best_result, best_km = apply_feature_cap(
        best_combined, {"pen": best_pen, **best_result}, best_km, U_wave, log_print=tprint
    )
    best_pen = best_result["pen"]

    out_doc = dict(source_doc)
    out_doc["pen"] = best_pen
    out_doc["mean_seg"] = best_result["mean_seg"]
    out_doc["U_wave"] = U_wave
    out_doc["start"] = model_start_ms
    out_doc["end"] = end_ms
    out_doc["kmeans_model"] = pickle.dumps(best_km)
    out_doc["feature_matrix"] = best_combined.tolist() if best_combined is not None else []
    save_to_state(mongo_db, imei, out_doc, collection_name=STATE_COLLECTION)
    return {
        "idx": idx, "outcome": "retrain_ok",
        "row": _build_excel_row(
            imei,
            mode,
            ("incremental_ok" if has_new_model else "first_train_ok"),
            old_mean_seg,
            {"mean_seg": best_result["mean_seg"]},
            old_range,
            current_range,
        ),
        "counters": {"retrain_ok": 1},
    }


def _merge_result_counters(counters_list):
    keys = ["retrain_ok", "seed_old_once", "keep_state"]
    out = {k: 0 for k in keys}
    for c in counters_list:
        for k in c:
            out[k] = out.get(k, 0) + c[k]
    return out


def main():
    run_started_at = time.time()
    run_end_ms = get_run_end_ms()
    output_file = resolve_output_filepath(run_end_ms)
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print("=" * 65)
    print("  device_model -> device_model_new scheduled retrain")
    print("  window={}d | workers={}".format(WINDOW_DAYS, WORKER_THREADS))
    print("  base(read)={}  state(write)={}  excel={}".format(
        BASE_COLLECTION, STATE_COLLECTION, output_file))
    print("  max_train_count={}".format(MAX_TRAIN_COUNT))
    print("  run_end={}".format(ts_to_str(run_end_ms)))
    print("=" * 65)

    print("\n> Connecting MongoDB...")
    try:
        mongo_client, mongo_db = _get_mongo_db()
        mongo_client.admin.command("ping")
        print("  [OK] MongoDB connected")
    except Exception as e:
        print("  [FAIL] {}".format(e))
        return

    print("\n> Loading device list from {} + {} ...".format(
        BASE_COLLECTION, STATE_COLLECTION))
    primary_list = load_primary_device_list(mongo_db)
    total_all = len(primary_list)
    if MAX_TRAIN_COUNT is not None and MAX_TRAIN_COUNT > 0:
        primary_list = primary_list[:MAX_TRAIN_COUNT]
    total = len(primary_list)
    print("  total union IMEIs: {}".format(total_all))
    print("  this run will process: {}\n".format(total))

    if not total:
        print("  [FAIL] no IMEIs found in {} or {}, exit".format(
            BASE_COLLECTION, STATE_COLLECTION))
        mongo_client.close()
        return

    print_lock = threading.Lock()

    def tprint(msg):
        with print_lock:
            print(msg, flush=True)

    work_items = [
        (idx, total, doc, mongo_db, tprint, run_end_ms)
        for idx, doc in enumerate(primary_list, 1)
    ]

    results = []
    with ThreadPoolExecutor(max_workers=WORKER_THREADS) as ex:
        for r in ex.map(process_device, work_items):
            results.append(r)

    results.sort(key=lambda x: x["idx"])
    mongo_client.close()

    counter_list = [r.get("counters", {}) for r in results]
    merged = _merge_result_counters(counter_list)

    rows = [r["row"] for r in results if r.get("row") is not None]

    summary = {
        "devices_processed": total,
        "retrain_ok": merged.get("retrain_ok", 0),
        "seed_old_once": merged.get("seed_old_once", 0),
        "keep_state": merged.get("keep_state", 0),
    }
    total_elapsed_seconds = round(time.time() - run_started_at, 2)
    summary["total_elapsed_seconds"] = total_elapsed_seconds
    summary["run_end"] = ts_to_str(run_end_ms)

    print("=" * 65)
    print("  done: {}".format(summary))
    print("=" * 65)

    if not rows:
        print("No rows to write to Excel (all skip / no row data).")
        return

    _write_excel(rows, output_file, summary)
    print("\n[OK] {} rows -> {}".format(len(rows), output_file))


# ============================================================
#  Excel
# ============================================================

_COL_ORDER = [
    "IMEI",
    "mode",
    "status",
    "old_train_range",
    "old_mean_seg",
    "new_train_range",
    "new_mean_seg",
]


def _write_excel(rows, filepath, summary):
    df   = pd.DataFrame(rows)
    cols = [c for c in _COL_ORDER if c in df.columns]
    df   = df[cols]

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="incremental", startrow=0)
        ws = writer.sheets["incremental"]

        hrow  = 1
        hfill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        for cell in ws[hrow]:
            cell.font = Font(bold=True)
            cell.fill = hfill

        for col_cells in ws.columns:
            max_len = max(
                (len(str(c.value)) for c in col_cells if c.value is not None),
                default=8,
            )
            ws.column_dimensions[
                get_column_letter(col_cells[0].column)
            ].width = min(max_len + 4, 36)

        summary_df = pd.DataFrame(
            [{"item": k, "value": v} for k, v in summary.items()]
        )
        summary_df.to_excel(writer, index=False, sheet_name="summary", startrow=0)
        ws2 = writer.sheets["summary"]
        for cell in ws2[1]:
            cell.font = Font(bold=True)
            cell.fill = hfill
        for col_cells in ws2.columns:
            max_len = max(
                (len(str(c.value)) for c in col_cells if c.value is not None),
                default=8,
            )
            ws2.column_dimensions[
                get_column_letter(col_cells[0].column)
            ].width = min(max_len + 4, 36)


# ============================================================
#  Scheduler
# ============================================================

def _append_log(log_file, message):
    os.makedirs(os.path.dirname(log_file), exist_ok=True)
    with open(log_file, "a", encoding="utf-8") as fh:
        fh.write(message.rstrip() + "\n")
    print(message, flush=True)


def run_scheduler():
    """Run one training batch on startup, then repeat every RETRAIN_INTERVAL_DAYS."""
    log_file = os.path.join(LOG_DIR, "incremental_train.log")
    interval_sec = RETRAIN_INTERVAL_DAYS * 24 * 60 * 60

    _append_log(log_file, "Incremental train scheduler started.")
    _append_log(log_file, "First run on startup, then every {} day(s).".format(RETRAIN_INTERVAL_DAYS))

    while True:
        _append_log(log_file, "=" * 60)
        _append_log(log_file, "Incremental train run at {}".format(
            datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S %z")))
        _append_log(log_file, "=" * 60)
        try:
            main()
        except Exception as exc:
            _append_log(log_file, "Run failed: {}".format(exc))
        _append_log(log_file, "Next incremental train in {} day(s).".format(RETRAIN_INTERVAL_DAYS))
        time.sleep(interval_sec)


def should_run_once():
    flag = os.environ.get("RUN_ONCE", "").strip().lower()
    return flag in ("1", "true", "yes")


# ============================================================
if __name__ == "__main__":
    if should_run_once():
        main()
    else:
        run_scheduler()
